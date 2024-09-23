#!/usr/bin/env python3

"""
Write Excel file containing iMessage data (user id, text, date, service, account)
Python 3.9+
"""

from datetime import datetime
import openpyxl
from openpyxl.styles import Font

class ExelWriter:
    """Handles exporting iMessage data to an Excel file."""

    def __init__(self, imessage_data: list, file_path: str):
        """Constructor method

        :param imessage_data: list with MessageData objects
                containing user id, text, date, service and account
        :param file_path: path to the location of the Excel file
        """
        self.imessage_data = imessage_data
        self.file_path = file_path

    def write_data(self):
        """Writes the iMessage data to an Excel file."""

        # Prepare data lists
        users = [data.user_id for data in self.imessage_data]
        messages = [data.text for data in self.imessage_data]
        dates = [data.date for data in self.imessage_data]
        services = [data.service for data in self.imessage_data]
        accounts = [data.account for data in self.imessage_data]
        is_from_me = [data.is_from_me for data in self.imessage_data]

        # Create and configure Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "iMessages"

        # Set header style
        bold16font = Font(size=16, bold=True)

        # Write headers
        headers = ["User ID", "Message", "Date", "Service", "Destination Caller ID", "Is From Me"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header).font = bold16font

        # Write data to columns
        for row, (user, message, date, service, account, from_me) in enumerate(zip(users, messages, dates, services, accounts, is_from_me), start=2):
            sheet.cell(row=row, column=1, value=user)
            sheet.cell(row=row, column=2, value=message)
            sheet.cell(row=row, column=3, value=date)
            sheet.cell(row=row, column=4, value=service)
            sheet.cell(row=row, column=5, value=account)
            sheet.cell(row=row, column=6, value=from_me)

        # Save the workbook (Excel file)
        try:
            file_name = f'iMessage-Data_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
            workbook.save(self.file_path + file_name)
            print("\n>>> Excel file successfully created! <<<")
            print(f"Find the file at: {self.file_path}{file_name}\n")
        except IOError as e:
            print("\n>>> Cannot write Excel file! <<<")
            print(e)